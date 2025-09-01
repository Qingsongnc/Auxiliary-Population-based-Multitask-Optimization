import scipy.stats as stats
import numpy as np
import json
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font

ForTasks = {
    'CEC17': [
        'CIHS', 'CIMS', 'CILS',
        'PIHS', 'PIMS', 'PILS',
        'NIHS', 'NIMS', 'NILS',
    ],
    'CEC22': [
        'Benchmark1', 'Benchmark2',
        'Benchmark3', 'Benchmark4',
        'Benchmark5', 'Benchmark6',
        'Benchmark7', 'Benchmark8',
        'Benchmark9', 'Benchmark10',
    ]
}
Comchar = {
    'win': '(+)',
    'tie': '(=)',
    'lose': '(-)',
}


centerAlign = Alignment(horizontal="center", vertical='center')
DefaultFont = Font(name='Times New Roman')
WinFont = Font(name='Times New Roman', bold=True)
TieFont = Font(name='Times New Roman', italic=True, underline='single')


def writeCell(sheet: Worksheet, row, col, Txt=None, font=None, Align=None):
    cell = sheet.cell(row, col, Txt)
    cell.alignment = centerAlign if Align is None else Align
    cell.font = DefaultFont if font is None else font
    cell.number_format = '@'

def SetStyle(sheet: Worksheet, row, col, Font=None, Alignment=None):
    cell = sheet.cell(row, col)
    cell.alignment = centerAlign if Alignment is None else Alignment
    cell.font = DefaultFont if Font is None else Font

def GetPos(startpos, algorithm, fnc, ColPerT, LinePerT, fmt='sign'):
    if fmt == 'number':
        return startpos + fnc + 1, 1 + algorithm
    return startpos + (ColPerT != 1) + 1 + int(fnc / ColPerT + 0.1), \
           2 + (LinePerT != 1) + algorithm * ColPerT + fnc % ColPerT


def Compare(A, B):
    if (A - B).any() == 0:
        return 'tie'
    else:
        res = stats.mannwhitneyu(A, B)
        if res.pvalue < 0.05:
            res = stats.mannwhitneyu(A, B, alternative='less')
            if res.pvalue < 0.05:
                return 'win', res.pvalue
            else:
                return 'lose', stats.mannwhitneyu(A, B, alternative='greater').pvalue
        else:
            return 'tie', res.pvalue


def outputTasks(sheet: Worksheet, Tasks, startpos=1, ColPerT=2, LinePerT=1, A=None, fmt='sign'):
    if fmt == 'number':
        if Tasks in ForTasks:
            Tasks = ForTasks[Tasks]
        else:
            assert A is not None, "When Tasks is not specified, A should be send here."
            Tasks = ['MTO' + str(i + 1) for i in range(int(A / ColPerT / LinePerT + 0.1))]
        for i in range(0, len(Tasks)):
            writeCell(sheet, startpos + 1 + i, 1, Tasks[i])
        writeCell(sheet, startpos + 1 + len(Tasks), 1, 'Number of "+/=/-"')
    else:
        assert not (ColPerT != 1 and LinePerT != 1), "The Multitask output is set on both lines and columns."
        writeCell(sheet, startpos, 1, Tasks)
        sheet.merge_cells(None, startpos, 1, startpos + (ColPerT != 1), 1 + (LinePerT != 1))
        if Tasks in ForTasks:
            Tasks = ForTasks[Tasks]
        else:
            assert A is not None, "When Tasks is not specified, A should be send here."
            Tasks = ['MTO' + str(i + 1) for i in range(int(A / ColPerT / LinePerT + 0.1))]
        for i in range(0, len(Tasks)):
            writeCell(sheet, startpos + 1 + (ColPerT != 1) + i * LinePerT, 1, Tasks[i])
            if LinePerT != 1:
                sheet.merge_cells(None, startpos + 1 + (ColPerT != 1) + i * LinePerT, 1,
                                  startpos + 1 + (ColPerT != 1) + i * LinePerT + LinePerT - 1, 1)
                for T in range(LinePerT):
                    Txt = 'T'
                    Txt = Txt + str(T + 1)
                    writeCell(sheet, startpos + 1 + (ColPerT != 1) + i * LinePerT + T, 2, Txt)
        writeCell(sheet, startpos + 1 + (ColPerT != 1) + len(Tasks) * LinePerT, 1,
                  'Number of "+/=/-"')
        sheet.merge_cells(None, startpos + 1 + (ColPerT != 1) + len(Tasks) * LinePerT, 1,
                          startpos + 1 + (ColPerT != 1) + len(Tasks) * LinePerT, 1 + (LinePerT != 1) + ColPerT)


def outputAlgorithm(sheet: Worksheet, Alg, AlgCom, startpos=1, ColPerT=2, LinePerT=1, fmt='sign'):
    assert not (ColPerT != 1 and LinePerT != 1), "The Multitask output is set on both lines and columns."
    if fmt == 'number':
        writeCell(sheet, startpos, 1, Alg.replace('_fit', ''))
        for i in range(len(AlgCom)):
            writeCell(sheet, startpos, 2 + i, AlgCom[i].replace('_fit', ''))
    else:
        writeCell(sheet, startpos, 2 + (LinePerT != 1), Alg.replace('_fit', ''))
        sheet.merge_cells(None, startpos, 2 + (LinePerT != 1), startpos, 2 + (LinePerT != 1) + ColPerT - 1)
        if ColPerT > 1:
            for T in range(ColPerT):
                Txt = 'T'
                if ColPerT != 1:
                    Txt += str(T + 1)
                writeCell(sheet, startpos + 1, 2 + (LinePerT != 1) + T, Txt)
        for i in range(len(AlgCom)):
            writeCell(sheet, startpos, 2 + (LinePerT != 1) + ColPerT + ColPerT * i, AlgCom[i].replace('_fit', ''))
            sheet.merge_cells(None, startpos, 2 + (LinePerT != 1) + ColPerT + ColPerT * i,
                              startpos, 2 + (LinePerT != 1) + ColPerT + ColPerT * i + ColPerT - 1)
            if ColPerT != 1:
                for T in range(ColPerT):
                    Txt = 'T'
                    if ColPerT != 1:
                        Txt += str(T + 1)
                    writeCell(sheet, startpos + 1, 2 + (LinePerT != 1) + ColPerT + ColPerT * i + T, Txt)


def outputData(sheet, A, Bs, startpos=1, ColPerT=1, LinePerT=2, fmt='sign'):
    if fmt == 'font':
        for i in range(A.shape[0]):
            pvalue = []
            ComList = []
            for j in range(len(Bs)):
                comchar, p = Compare(A[i], Bs[j][i])
                ComList.append(comchar)
                pvalue.append(p)
            winone = 0
            row, col = GetPos(startpos, 0, i, ColPerT, LinePerT)
            if 'lose' in ComList:
                winone = 1
                for j in range(len(ComList)):
                    if ComList[j] == 'lose':
                        if Compare(Bs[j][i], Bs[winone-1][i]) == 'win':
                            winone = j + 1
                        elif Compare(Bs[j][i], Bs[winone-1][i]) == 'tie':
                            if Bs[j][i].mean() < Bs[winone-1].mean():
                                winone = j + 1
                writeCell(sheet, row, col, "{:.2e}\n{:.2e}\n-".format(A[i].mean(), A[i].std()))
            elif 'tie' in ComList:
                bestList = A[i]
                for j in range(len(ComList)):
                    if ComList[j] == 'tie':
                        if Bs[j][i].mean() < bestList.mean():
                            bestList = Bs[j][i]
                            winone = j + 1
                if winone == 0:
                    writeCell(sheet, row, col, "{:.2e}\n{:.2e}\n-".format(A[i].mean(), A[i].std()), WinFont)
                else:
                    writeCell(sheet, row, col, "{:.2e}\n{:.2e}\n-".format(A[i].mean(), A[i].std()))
            else:
                writeCell(sheet, row, col, "{:.2e}\n{:.2e}\n-".format(A[i].mean(), A[i].std()), WinFont)
            for j in range(len(Bs)):
                row, col = GetPos(startpos, j + 1, i, ColPerT, LinePerT)
                if j + 1 == winone:
                    writeCell(sheet, row, col, "{:.2e}{}\n{:.2e}\n{:.4f}".format(Bs[j][i].mean(), Comchar[ComList[j]], A[i].std(), pvalue[j]), WinFont)
                else:
                    writeCell(sheet, row, col, "{:.2e}{}\n{:.2e}\n{:.4f}".format(Bs[j][i].mean(), Comchar[ComList[j]], A[i].std(), pvalue[j]))
                writeCell(sheet, row, col + ColPerT * len(Bs), str(ComList[j]))
    elif fmt == 'sign':
        for i in range(A.shape[0]):
            ComList = []
            for j in range(len(Bs)):
                ComList.append(Compare(A[i], Bs[j][i]))
            row, col = GetPos(startpos, 0, i, ColPerT, LinePerT)
            writeCell(sheet, row, col, "{:.2e}".format(A[i].mean()))
            for j in range(len(Bs)):
                row, col = GetPos(startpos, j + 1, i, ColPerT, LinePerT)
                writeCell(sheet, row, col, "{:.2e}{}".format(Bs[j][i].mean(), Comchar[ComList[j]]))
                writeCell(sheet, row, col + ColPerT * len(Bs), str(ComList[j]))
    elif fmt == 'number':
        for i in range(int(A.shape[0] / (ColPerT * LinePerT) + 0.1)):
            ComList = []
            for j in range(len(Bs)):
                win = 0
                tie = 0
                lose = 0
                for k in range(ColPerT * LinePerT):
                    Comch = Compare(A[i * ColPerT * LinePerT + k], Bs[j][i * ColPerT * LinePerT + k])
                    if Comch == 'win':
                        win += 1
                    elif Comch == 'tie':
                        tie += 1
                    else:
                        lose += 1
                if -5 <= win - lose <= 5:
                    ComList.append('tie')
                elif win - lose > 5:
                    ComList.append('win')
                else:
                    ComList.append('lose')
                row, col = GetPos(startpos, j + 1, i, ColPerT, LinePerT, fmt)
                writeCell(sheet, row, col, "{}/{}/{}{}".format(win, tie, lose, Comchar[ComList[j]]))
                writeCell(sheet, row, col + len(Bs), str(ComList[j]))
    # Output win/tie/lose words right side of the table data.
    if fmt == 'number':
        for i in range(len(Bs)):
            rows = startpos + 2 + int(len(A) / (ColPerT * LinePerT + 0.1))
            cols = 2
            ComLists = sheet.iter_cols(cols + len(Bs), cols + len(Bs) * 2 - 1, values_only=True)
            for j, ComList in enumerate(ComLists):
                comlist = [com for com in ComList]
                comlist = comlist[startpos:]
                win = len([com for com in comlist if com == 'win'])
                tie = len([com for com in comlist if com == 'tie'])
                lose = len([com for com in comlist if com == 'lose'])
                writeCell(sheet, rows, cols + j, "{}/{}/{}".format(win, tie, lose))
    else:
        for i in range(len(Bs) * ColPerT):
            rows = startpos + (ColPerT != 1) + 1 + int(len(A) / ColPerT + 0.1)
            cols = 2 + (LinePerT != 1) + ColPerT
            ComLists = sheet.iter_cols(cols + len(Bs) * ColPerT, cols + len(Bs) * ColPerT * 2 - 1, values_only=True)
            for j, ComList in enumerate(ComLists):
                comlist = [com for com in ComList]
                comlist = comlist[startpos:]
                win = len([com for com in comlist if com == 'win'])
                tie = len([com for com in comlist if com == 'tie'])
                lose = len([com for com in comlist if com == 'lose'])
                writeCell(sheet, rows, cols + j, "{}/{}/{}".format(win, tie, lose))


def outputSet(sheet, Tasks, A, Bs, filename1, filenames2, startpos=1, ColPerT=1, LinePerT=2, fmt='sign'):
    outputTasks(sheet, Tasks, startpos=startpos, ColPerT=ColPerT, LinePerT=LinePerT, A=A.shape[0], fmt=fmt)
    outputAlgorithm(sheet, filename1, filenames2, startpos=startpos, ColPerT=ColPerT, LinePerT=LinePerT, fmt=fmt)
    outputData(sheet, A, Bs, startpos, ColPerT, LinePerT, fmt)


def readFile(filename):
    file1 = open('source/' + filename + '.csv', 'r')
    A = file1.readlines()
    A = [line.replace(',\n', '').split(',') for line in A]
    A = np.array(A, dtype=float).T
    A[A < 0] = 0
    return A


def main(filename1, filenames2, algtype='17+22', T=2, LinePerT=1, outputFile=None, fmt='sign', BestBold=False):
    # Create workbook
    if outputFile is None:
        outputFile = filename1
    outputFile = 'output/output' + outputFile + '.xlsx'
    ws = Workbook()
    sheet = ws.active
    # Read The Algorithm Files
    A = readFile(filename1)
    Bs = []
    for filename in filenames2:
        Bs.append(readFile(filename))
        assert len(A) == len(Bs[-1]), "File contents are not Paired. A.shape=({},{}) but B.shape=({},{})".format(
            A.shape[0], A.shape[1], Bs[-1].shape[0], Bs[-1].shape[1])

    # Output Titles for algorithms and tasks
    if algtype == '17+22':
        if T == 1:
            LinePerT = int(A.shape[0] / (len(ForTasks['CEC17']) + len(ForTasks['CEC22'])) + 0.1)
        # For CEC17
        outputSet(sheet, 'CEC17', A[:len(ForTasks['CEC17']) * T * LinePerT],
                  [B[:len(ForTasks['CEC17']) * T * LinePerT] for B in Bs], filename1, filenames2,
                  ColPerT=T, LinePerT=LinePerT, fmt=fmt)
        # For CEC22
        startpos = len(ForTasks['CEC17']) * LinePerT + 5
        outputSet(sheet, 'CEC22', A[len(ForTasks['CEC17']) * T * LinePerT:],
                  [B[len(ForTasks['CEC17']) * T * LinePerT:] for B in Bs], filename1, filenames2,
                  startpos=startpos, ColPerT=T, LinePerT=LinePerT, fmt=fmt)
    elif algtype == '17':
        if T == 1:
            LinePerT = int(A.shape[0] / len(ForTasks['CEC17']) + 0.1)
        outputSet(sheet, 'CEC17', A, Bs, filename1, filenames2,
                  ColPerT=T, LinePerT=LinePerT, fmt=fmt)
    elif algtype == '22':
        if T == 1:
            LinePerT = int(A.shape[0] / len(ForTasks['CEC22']) + 0.1)
        outputSet(sheet, 'CEC22', A, Bs, filename1, filenames2,
                  ColPerT=T, LinePerT=LinePerT, fmt=fmt)
    else:
        outputSet(sheet, 'MTO', A, Bs, filename1, filenames2,
                  ColPerT=T, LinePerT=LinePerT, fmt=fmt)
    # Save File
    ws.save(outputFile)


if __name__ == '__main__':
    config = open('config.json', 'r')
    config = json.load(config)
    if 'Tasks' not in config:
        tasks = 'None'
    else:
        tasks = config['Tasks']
    if 'CompareSel' in config:
        keySel = config['CompareSel']
    else:
        keySel = 'Compare'
    if 'fmt' in config:
        fmt = config['fmt']
    else:
        fmt = 'sign'
    if 'ColPerT' in config:
        ColPerT = config['ColPerT']
    else:
        ColPerT = 1
    if 'LinePerT' in config:
        LinePerT = config['LinePerT']
    else:
        LinePerT = 2
    A = config['Algorithm']
    B = config[keySel]
    outputFile = config['output']
    # A file tends to be less than files in B
    main(A, B, tasks, T=ColPerT, LinePerT=LinePerT, outputFile=outputFile, fmt=fmt)
